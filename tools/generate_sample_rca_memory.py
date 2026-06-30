"""Generate a deterministic sample past-RCA memory workbook for demos."""

from __future__ import annotations

import argparse
import random
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


COLUMNS = [
    "incident_id",
    "date",
    "system_area",
    "service_name",
    "error_signature",
    "problem_statement",
    "symptoms",
    "root_cause",
    "immediate_fix",
    "long_term_fix",
    "evidence_checked",
    "owner_team",
    "tags",
    "confidence",
    "status",
]

DEFAULT_ROWS = 512
DEFAULT_OUTPUT = Path("data") / "past_rca_memory_sample.xlsx"
START_DATE = date(2025, 1, 6)


CURATED_INCIDENTS: list[dict[str, Any]] = [
    {
        "system_area": "Identity & Access",
        "service_name": "login-api",
        "error_signature": "HTTP 500 /auth/token after deployment",
        "problem_statement": "Login API returns 500 after the latest deployment.",
        "symptoms": "Users cannot sign in; token endpoint error rate above 35%; rollback restores service.",
        "root_cause": "Deployment introduced a missing JWT issuer environment variable in the login-api runtime configuration.",
        "immediate_fix": "Restore the JWT issuer variable and restart the login-api pods.",
        "long_term_fix": "Add deployment-time config validation and a canary check for token issuance.",
        "evidence_checked": "deployment diff; pod env dump; auth service logs; rollback comparison",
        "owner_team": "Identity Platform",
        "tags": "auth, login, deployment, environment, jwt, known-fix",
        "confidence": "high",
        "status": "resolved",
    },
    {
        "system_area": "Network",
        "service_name": "vpn-gateway",
        "error_signature": "VPN handshake timeout after firewall rule update",
        "problem_statement": "VPN access failures started after firewall changes.",
        "symptoms": "Employees cannot connect to VPN; SSL handshake times out; office network still works.",
        "root_cause": "A new firewall rule blocked UDP 4500 traffic required for VPN tunnel negotiation.",
        "immediate_fix": "Reopen UDP 4500 from approved employee IP ranges.",
        "long_term_fix": "Add firewall change validation against the VPN required-port checklist.",
        "evidence_checked": "firewall change log; VPN gateway logs; packet capture; access rule diff",
        "owner_team": "Network Operations",
        "tags": "vpn, firewall, network, udp, access, known-fix",
        "confidence": "high",
        "status": "resolved",
    },
    {
        "system_area": "Payments",
        "service_name": "payment-api",
        "error_signature": "Payment timeout after database pool change",
        "problem_statement": "Payment requests time out after database pool configuration changes.",
        "symptoms": "Checkout stalls; payment API p95 latency exceeds 20 seconds; database wait time spikes.",
        "root_cause": "The payment-api database connection pool was reduced below peak checkout concurrency.",
        "immediate_fix": "Raise the payment-api pool limit and recycle affected workers.",
        "long_term_fix": "Load-test pool settings and add alerts for connection wait saturation.",
        "evidence_checked": "database pool metrics; payment-api latency traces; config commit; checkout logs",
        "owner_team": "Payments",
        "tags": "payments, checkout, database, pool, timeout, known-fix",
        "confidence": "high",
        "status": "resolved",
    },
    {
        "system_area": "Batch Processing",
        "service_name": "invoice-worker",
        "error_signature": "Invoice batch failed after schema migration",
        "problem_statement": "Nightly invoice job fails after a schema migration.",
        "symptoms": "Invoice batch exits early; new invoices not generated; worker logs show missing column.",
        "root_cause": "The invoice worker still queried the deprecated invoice_status column after migration.",
        "immediate_fix": "Deploy the worker patch that reads billing_status and rerun the failed batch.",
        "long_term_fix": "Add migration compatibility tests for all scheduled billing jobs.",
        "evidence_checked": "migration SQL; worker stack trace; failed job logs; schema inspection",
        "owner_team": "Billing Platform",
        "tags": "billing, invoice, batch, schema, migration, known-fix",
        "confidence": "high",
        "status": "resolved",
    },
    {
        "system_area": "Observability",
        "service_name": "alert-router",
        "error_signature": "Monitoring alert caused by stale service configuration",
        "problem_statement": "Monitoring alerts fire repeatedly after a stale service configuration is deployed.",
        "symptoms": "False critical alerts; alert-router routes old service name; on-call receives duplicate pages.",
        "root_cause": "The alert-router config map still referenced a retired service identifier.",
        "immediate_fix": "Update the config map and reload alert-router.",
        "long_term_fix": "Validate alert routes against the service catalog during deployment.",
        "evidence_checked": "alert payload; config map; service catalog; deployment event timeline",
        "owner_team": "SRE Observability",
        "tags": "monitoring, alerting, stale-config, service-catalog, known-fix",
        "confidence": "high",
        "status": "resolved",
    },
    {
        "system_area": "Commerce",
        "service_name": "checkout-api",
        "error_signature": "Checkout timeout after database migration",
        "problem_statement": "Checkout requests time out after a customer table database migration.",
        "symptoms": "Checkout API times out; database CPU rises; slow query log shows missing index usage.",
        "root_cause": "The migration dropped the covering index used by checkout customer lookups.",
        "immediate_fix": "Recreate the customer lookup index and clear stuck checkout requests.",
        "long_term_fix": "Add query-plan checks to migration review for checkout-critical tables.",
        "evidence_checked": "slow query log; migration diff; query plan; database index list",
        "owner_team": "Commerce Platform",
        "tags": "checkout, database, migration, index, latency, known-fix",
        "confidence": "high",
        "status": "resolved",
    },
]


SCENARIO_FAMILIES: list[dict[str, Any]] = [
    {
        "system_area": "Identity & Access",
        "service_name": "login-api",
        "owner_team": "Identity Platform",
        "error_signatures": [
            "HTTP 500 /auth/token after deployment",
            "TokenClaimsMapper null reference",
            "OIDC callback rejected issuer mismatch",
            "Session refresh returns invalid_grant",
        ],
        "problem_templates": [
            "Login API returns 500 after {change}.",
            "Users cannot authenticate after {change}.",
            "Session refresh fails after {change}.",
        ],
        "changes": [
            "a deployment",
            "config promotion",
            "identity provider metadata refresh",
            "secret rotation",
        ],
        "symptoms": [
            "users cannot sign in",
            "token endpoint error rate increased",
            "rollback restores authentication",
            "OIDC callback logs show validation failure",
        ],
        "root_causes": [
            "missing JWT issuer environment variable",
            "stale OIDC discovery metadata in cache",
            "rotated client secret not deployed to one environment",
            "audience claim validation configured for the wrong tenant",
        ],
        "immediate_fixes": [
            "restore the correct auth environment variable and restart pods",
            "refresh OIDC metadata cache and restart the login workers",
            "sync the rotated client secret to the affected environment",
            "correct the tenant audience setting and redeploy login-api",
        ],
        "long_term_fixes": [
            "add deployment-time auth config validation",
            "add a synthetic login canary to release gates",
            "automate secret rollout checks across environments",
            "add tenant-claim contract tests for identity changes",
        ],
        "evidence": [
            "deployment diff",
            "pod env dump",
            "auth service logs",
            "OIDC metadata response",
            "rollback comparison",
        ],
        "tags": ["auth", "login", "deployment", "jwt", "oidc"],
    },
    {
        "system_area": "Payments",
        "service_name": "payment-api",
        "owner_team": "Payments",
        "error_signatures": [
            "Payment timeout after database pool change",
            "Gateway 504 from payment authorization",
            "Connection wait exceeded pool timeout",
            "Payment capture retries exhausted",
        ],
        "problem_templates": [
            "Payment requests time out after {change}.",
            "Payment authorization latency spikes after {change}.",
            "Checkout payment capture fails after {change}.",
        ],
        "changes": [
            "database pool tuning",
            "gateway SDK upgrade",
            "checkout release",
            "read replica failover",
        ],
        "symptoms": [
            "checkout stalls",
            "payment p95 latency exceeds threshold",
            "database connection wait time spikes",
            "retry queue grows quickly",
        ],
        "root_causes": [
            "database connection pool reduced below peak checkout concurrency",
            "gateway timeout setting lowered during SDK upgrade",
            "read replica endpoint configured for write traffic",
            "idempotency key store latency caused duplicate capture retries",
        ],
        "immediate_fixes": [
            "raise the payment-api pool limit and recycle workers",
            "restore the gateway timeout setting",
            "point payment writes back to the primary database endpoint",
            "pause duplicate capture retries and drain the retry queue",
        ],
        "long_term_fixes": [
            "load-test payment pool settings before production changes",
            "add alerts for connection wait saturation",
            "add release checks for payment gateway timeout defaults",
            "add idempotency-store latency SLOs",
        ],
        "evidence": [
            "database pool metrics",
            "payment traces",
            "gateway logs",
            "config commit",
            "checkout error dashboard",
        ],
        "tags": ["payments", "checkout", "database", "timeout", "pool"],
    },
    {
        "system_area": "Network",
        "service_name": "vpn-gateway",
        "owner_team": "Network Operations",
        "error_signatures": [
            "VPN handshake timeout after firewall rule update",
            "SAML VPN login reaches callback but tunnel fails",
            "IKE negotiation failed no proposal chosen",
            "Split tunnel routes missing for internal CIDR",
        ],
        "problem_templates": [
            "VPN access failures started after {change}.",
            "Employees cannot connect to VPN after {change}.",
            "Internal apps are unreachable over VPN after {change}.",
        ],
        "changes": [
            "firewall rule cleanup",
            "VPN gateway upgrade",
            "network ACL update",
            "route table change",
        ],
        "symptoms": [
            "employees cannot connect to VPN",
            "handshake times out",
            "internal service routes are unreachable",
            "office network access remains healthy",
        ],
        "root_causes": [
            "firewall rule blocked UDP 4500 for VPN negotiation",
            "route table no longer advertised the internal service CIDR",
            "new gateway cipher set did not match client policy",
            "network ACL blocked return traffic from the VPN subnet",
        ],
        "immediate_fixes": [
            "reopen UDP 4500 from approved employee IP ranges",
            "restore the internal CIDR route advertisement",
            "enable the previous cipher suite during client rollout",
            "allow return traffic from the VPN subnet in the ACL",
        ],
        "long_term_fixes": [
            "add firewall validation against VPN required ports",
            "test route reachability before network changes",
            "publish VPN client policy before gateway cipher updates",
            "add automated VPN smoke tests after ACL changes",
        ],
        "evidence": [
            "firewall change log",
            "VPN gateway logs",
            "packet capture",
            "route table diff",
            "client connection logs",
        ],
        "tags": ["vpn", "firewall", "network", "access", "route"],
    },
    {
        "system_area": "Billing",
        "service_name": "invoice-worker",
        "owner_team": "Billing Platform",
        "error_signatures": [
            "Invoice batch failed after schema migration",
            "Scheduler skipped invoice generation window",
            "Invoice PDF renderer out of memory",
            "Tax calculation job returns validation error",
        ],
        "problem_templates": [
            "Invoice jobs stopped after {change}.",
            "Nightly invoice batch fails after {change}.",
            "Billing reports are delayed after {change}.",
        ],
        "changes": [
            "schema migration",
            "scheduler change",
            "tax rules update",
            "worker image release",
        ],
        "symptoms": [
            "new invoices are not generated",
            "worker exits early",
            "scheduler shows missed run window",
            "billing queue backlog grows",
        ],
        "root_causes": [
            "worker still queried a deprecated invoice status column",
            "scheduler timezone changed from UTC to local time",
            "tax rule file was deployed without the required region mapping",
            "PDF renderer memory limit was reduced below peak invoice size",
        ],
        "immediate_fixes": [
            "deploy the worker patch and rerun failed batches",
            "restore UTC scheduler configuration and trigger the missed run",
            "restore the region mapping file and rerun tax validation",
            "raise renderer memory limit and resume invoice generation",
        ],
        "long_term_fixes": [
            "add migration compatibility tests for scheduled billing jobs",
            "lock scheduler timezone in deployment policy",
            "validate tax rule files before release",
            "add invoice PDF load tests to worker release checks",
        ],
        "evidence": [
            "migration SQL",
            "worker stack trace",
            "scheduler history",
            "billing queue metrics",
            "tax validation report",
        ],
        "tags": ["billing", "invoice", "batch", "schema", "scheduler"],
    },
    {
        "system_area": "Data Platform",
        "service_name": "orders-etl",
        "owner_team": "Data Engineering",
        "error_signatures": [
            "ETL freshness breach after source schema drift",
            "Orders pipeline stuck in transform stage",
            "Warehouse load failed duplicate key",
            "Partition discovery skipped current day",
        ],
        "problem_templates": [
            "Orders data freshness breached after {change}.",
            "ETL pipeline fails after {change}.",
            "Warehouse order table is stale after {change}.",
        ],
        "changes": [
            "upstream schema change",
            "partition configuration update",
            "warehouse loader release",
            "source connector upgrade",
        ],
        "symptoms": [
            "dashboard data is stale",
            "transform job retries repeatedly",
            "warehouse load duration increases",
            "freshness alert fires",
        ],
        "root_causes": [
            "source added a nullable column not handled by the transform contract",
            "partition path pattern no longer matched current-day files",
            "loader deduplication key did not include region",
            "connector upgrade changed timestamp parsing behavior",
        ],
        "immediate_fixes": [
            "patch the transform contract and backfill the missed window",
            "restore the partition path pattern and rerun discovery",
            "update deduplication key and reload affected partitions",
            "pin timestamp parsing mode and replay the connector batch",
        ],
        "long_term_fixes": [
            "add schema-contract checks for upstream changes",
            "test partition discovery in pre-prod",
            "add duplicate-key detection before warehouse load",
            "add connector upgrade replay tests",
        ],
        "evidence": [
            "ETL job logs",
            "schema registry diff",
            "warehouse load history",
            "freshness dashboard",
            "partition listing",
        ],
        "tags": ["etl", "warehouse", "schema", "freshness", "orders"],
    },
    {
        "system_area": "Customer Portal",
        "service_name": "portal-web",
        "owner_team": "Experience Platform",
        "error_signatures": [
            "Blank page after frontend deployment",
            "Static asset 404 due CDN cache mismatch",
            "Feature flag hides support widget",
            "Browser console reports chunk load error",
        ],
        "problem_templates": [
            "Customer portal shows a blank page after {change}.",
            "Portal assets fail to load after {change}.",
            "Support widget disappears after {change}.",
        ],
        "changes": [
            "frontend deployment",
            "CDN purge",
            "feature flag rollout",
            "asset pipeline change",
        ],
        "symptoms": [
            "customers see a blank page",
            "static assets return 404",
            "console shows chunk load errors",
            "support widget is missing",
        ],
        "root_causes": [
            "CDN retained old HTML pointing to removed JavaScript chunks",
            "asset manifest was published after the release switched traffic",
            "feature flag targeted all customers instead of beta users",
            "base path changed without matching router configuration",
        ],
        "immediate_fixes": [
            "purge CDN HTML and restore the previous asset manifest",
            "republish the asset manifest and restart the release",
            "limit the feature flag to beta users",
            "restore the router base path and redeploy portal-web",
        ],
        "long_term_fixes": [
            "make frontend releases atomic across HTML and assets",
            "add CDN cache validation to release checks",
            "require feature-flag blast-radius review",
            "add browser smoke tests for routed pages",
        ],
        "evidence": [
            "browser console",
            "CDN logs",
            "asset manifest diff",
            "feature flag audit",
            "release timeline",
        ],
        "tags": ["frontend", "cdn", "portal", "assets", "feature-flag"],
    },
    {
        "system_area": "Notifications",
        "service_name": "email-service",
        "owner_team": "Messaging Platform",
        "error_signatures": [
            "Email delivery failure after SMTP credential rotation",
            "SMS webhook signature validation failed",
            "Notification queue retry storm",
            "Push notification provider returns 401",
        ],
        "problem_templates": [
            "Notifications fail after {change}.",
            "Email delivery drops after {change}.",
            "Notification queue backs up after {change}.",
        ],
        "changes": [
            "credential rotation",
            "provider API update",
            "queue worker deployment",
            "template release",
        ],
        "symptoms": [
            "customer emails are delayed",
            "provider returns authentication errors",
            "notification queue depth grows",
            "retry count spikes",
        ],
        "root_causes": [
            "rotated SMTP password was not synced to the email-service secret",
            "provider webhook secret was updated in only one environment",
            "queue worker retry interval was lowered too aggressively",
            "template release introduced invalid personalization variables",
        ],
        "immediate_fixes": [
            "sync the SMTP password and restart email-service",
            "update the webhook secret in all environments",
            "restore retry interval and drain delayed messages",
            "rollback the invalid template and replay failed notifications",
        ],
        "long_term_fixes": [
            "add secret sync verification after rotations",
            "add provider webhook contract tests",
            "bound retry rate with queue backpressure",
            "validate template variables before release",
        ],
        "evidence": [
            "provider response logs",
            "secret version history",
            "queue metrics",
            "template diff",
            "delivery dashboard",
        ],
        "tags": ["notifications", "email", "secret", "queue", "provider"],
    },
    {
        "system_area": "Inventory",
        "service_name": "stock-sync",
        "owner_team": "Supply Chain Systems",
        "error_signatures": [
            "Inventory sync lag after message broker change",
            "Stock counts mismatch after retry replay",
            "Warehouse feed parser rejects SKU batch",
            "Consumer lag rises after partition reassignment",
        ],
        "problem_templates": [
            "Inventory updates lag after {change}.",
            "Stock counts mismatch after {change}.",
            "Warehouse feed ingestion fails after {change}.",
        ],
        "changes": [
            "broker partition reassignment",
            "warehouse feed format update",
            "stock-sync deployment",
            "retry replay",
        ],
        "symptoms": [
            "stock counts are stale",
            "consumer lag rises",
            "warehouse feed rejects SKU rows",
            "manual reconciliation finds mismatches",
        ],
        "root_causes": [
            "stock-sync consumers were pinned to old partition assignments",
            "warehouse feed added a mandatory location code field",
            "retry replay processed non-idempotent inventory adjustments",
            "SKU parser rejected lowercase warehouse identifiers",
        ],
        "immediate_fixes": [
            "restart consumers and rebalance partition assignments",
            "update parser mapping and replay the rejected feed",
            "halt replay and reverse duplicate inventory adjustments",
            "normalize warehouse identifiers before SKU validation",
        ],
        "long_term_fixes": [
            "automate broker rebalance checks after partition changes",
            "add feed schema validation with warehouse partners",
            "make inventory adjustments idempotent",
            "add parser tests for warehouse identifier variants",
        ],
        "evidence": [
            "consumer group metrics",
            "warehouse feed sample",
            "stock adjustment audit",
            "parser error logs",
            "broker reassignment plan",
        ],
        "tags": ["inventory", "broker", "consumer-lag", "warehouse", "sync"],
    },
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--rows", type=int, default=DEFAULT_ROWS, help="Number of rows to generate.")
    parser.add_argument("--seed", type=int, default=42, help="Random seed for deterministic output.")
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Output .xlsx path.",
    )
    return parser.parse_args()


def _join_sample(values: list[str], rng: random.Random, count: int) -> str:
    return "; ".join(rng.sample(values, k=min(count, len(values))))


def _choice(values: list[str], rng: random.Random) -> str:
    return values[rng.randrange(len(values))]


def build_rows(row_count: int, seed: int) -> list[dict[str, Any]]:
    if row_count < len(CURATED_INCIDENTS):
        raise ValueError(f"rows must be at least {len(CURATED_INCIDENTS)}")
    if row_count > 10_000:
        raise ValueError("rows must be 10,000 or fewer for this sample generator")

    rng = random.Random(seed)
    rows: list[dict[str, Any]] = []

    for index, incident in enumerate(CURATED_INCIDENTS, start=1):
        incident_date = START_DATE + timedelta(days=index * 9)
        rows.append(
            {
                "incident_id": f"RCA-{incident_date:%Y}-{index:04d}",
                "date": incident_date,
                **incident,
            }
        )

    statuses = ["resolved", "resolved", "resolved", "monitoring", "open"]
    confidences = ["high", "high", "medium", "medium", "low"]

    for index in range(len(rows) + 1, row_count + 1):
        family = SCENARIO_FAMILIES[(index - 1) % len(SCENARIO_FAMILIES)]
        incident_date = START_DATE + timedelta(days=(index * 3) % 520)
        change = _choice(family["changes"], rng)
        root_cause = _choice(family["root_causes"], rng)
        immediate_fix = _choice(family["immediate_fixes"], rng)
        long_term_fix = _choice(family["long_term_fixes"], rng)
        problem_template = _choice(family["problem_templates"], rng)

        tags = list(family["tags"])
        if "known-fix" not in tags and rng.random() < 0.58:
            tags.append("known-fix")
        if rng.random() < 0.22:
            tags.append("repeat-incident")

        rows.append(
            {
                "incident_id": f"RCA-{incident_date:%Y}-{index:04d}",
                "date": incident_date,
                "system_area": family["system_area"],
                "service_name": family["service_name"],
                "error_signature": _choice(family["error_signatures"], rng),
                "problem_statement": problem_template.format(change=change),
                "symptoms": _join_sample(family["symptoms"], rng, 3),
                "root_cause": root_cause,
                "immediate_fix": immediate_fix,
                "long_term_fix": long_term_fix,
                "evidence_checked": _join_sample(family["evidence"], rng, 4),
                "owner_team": family["owner_team"],
                "tags": ", ".join(tags),
                "confidence": _choice(confidences, rng),
                "status": _choice(statuses, rng),
            }
        )

    return rows


def build_workbook(rows: int, seed: int, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(build_rows(rows, seed), columns=COLUMNS)
    df["date"] = pd.to_datetime(df["date"])

    with pd.ExcelWriter(output_path, engine="openpyxl", date_format="yyyy-mm-dd") as writer:
        df.to_excel(writer, index=False, sheet_name="Past RCA Memory")
        summary_rows = [
            ["Purpose", "Sample read-only past RCA memory for local demo retrieval."],
            ["Rows", rows],
            ["Generated By", "tools/generate_sample_rca_memory.py"],
            ["Schema", ", ".join(COLUMNS)],
            ["Safety", "Dummy data only. Similar incidents are hints, not final truth."],
        ]
        pd.DataFrame(summary_rows, columns=["Field", "Value"]).to_excel(
            writer, index=False, sheet_name="README"
        )

    apply_formatting(output_path, rows)
    return output_path


def apply_formatting(path: Path, rows: int) -> None:
    workbook = load_workbook(path)
    memory = workbook["Past RCA Memory"]
    readme = workbook["README"]

    header_fill = PatternFill("solid", fgColor="123047")
    header_font = Font(color="FFFFFF", bold=True)
    subtle_fill = PatternFill("solid", fgColor="EAF2F8")
    border = Border(bottom=Side(style="thin", color="D6DEE6"))

    memory.freeze_panes = "A2"
    memory.sheet_view.showGridLines = False

    for cell in memory[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    widths = {
        "A": 16,
        "B": 12,
        "C": 22,
        "D": 18,
        "E": 34,
        "F": 44,
        "G": 48,
        "H": 52,
        "I": 42,
        "J": 48,
        "K": 42,
        "L": 24,
        "M": 34,
        "N": 12,
        "O": 14,
    }
    for column, width in widths.items():
        memory.column_dimensions[column].width = width

    for row in memory.iter_rows(min_row=2, max_row=rows + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        row[1].number_format = "yyyy-mm-dd"

    confidence_validation = DataValidation(
        type="list",
        formula1='"high,medium,low"',
        allow_blank=False,
    )
    status_validation = DataValidation(
        type="list",
        formula1='"resolved,monitoring,open"',
        allow_blank=False,
    )
    memory.add_data_validation(confidence_validation)
    memory.add_data_validation(status_validation)
    confidence_validation.add(f"N2:N{rows + 1}")
    status_validation.add(f"O2:O{rows + 1}")

    memory.conditional_formatting.add(
        f"O2:O{rows + 1}",
        FormulaRule(
            formula=['$O2="open"'],
            fill=PatternFill("solid", fgColor="FCE4D6"),
        ),
    )
    memory.conditional_formatting.add(
        f"O2:O{rows + 1}",
        FormulaRule(
            formula=['$O2="monitoring"'],
            fill=PatternFill("solid", fgColor="FFF2CC"),
        ),
    )
    memory.conditional_formatting.add(
        f"N2:N{rows + 1}",
        FormulaRule(
            formula=['$N2="high"'],
            fill=PatternFill("solid", fgColor="E2F0D9"),
        ),
    )

    readme.sheet_view.showGridLines = False
    readme.column_dimensions["A"].width = 20
    readme.column_dimensions["B"].width = 110
    for cell in readme[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in readme.iter_rows(min_row=2, max_row=readme.max_row):
        row[0].fill = subtle_fill
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(path)


def main() -> None:
    args = parse_args()
    output = build_workbook(args.rows, args.seed, args.output)
    print(output.resolve())


if __name__ == "__main__":
    main()
