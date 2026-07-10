from __future__ import annotations

from typing import Any

from conftest import CapturingStubProvider
from openpyxl import load_workbook

from agent.orchestrator import RCAAgent
from config import Settings
from memory import MemorySearch
from providers.base import RCAProvider
from schemas import RCAInput, RCAReport


def report_payload(root_cause: str, selected_cause: str) -> dict[str, Any]:
    return {
        "problem": "Checkout requests time out after a database migration.",
        "summary": "The incident follows from a weak migration validation process.",
        "why_chain": [
            {
                "index": 1,
                "question": "Why did checkout requests time out?",
                "answer": "Checkout database queries became too slow after the migration.",
            },
            {
                "index": 2,
                "question": "Why did the database queries become too slow?",
                "answer": "The migration changed the checkout query plan unexpectedly.",
            },
            {
                "index": 3,
                "question": "Why was that query-plan change released?",
                "answer": "The migration release path lacked a performance validation gate.",
            },
        ],
        "root_cause": root_cause,
        "contributing_factors": [
            "Query-plan review was not required for migrations.",
            "Slow-query alerting was not connected to checkout impact.",
        ],
        "recommendations": [
            "Add a performance validation gate to database migrations.",
            "Require query-plan diffs for checkout-critical migrations.",
        ],
        "assumptions": [],
        "evidence_needed": ["Migration query-plan diff", "Checkout slow-query logs"],
        "validation_notes": [],
        "method_detail": {
            "fishbone": {
                "categories": {
                    "People": ["Migration review ownership was unclear."],
                    "Process": [selected_cause],
                    "Tooling": [],
                    "Environment": [],
                    "Data": [],
                },
                "selected_category": "Process",
                "selected_cause": selected_cause,
            }
        },
        "confidence": "medium",
    }


class StubProvider(RCAProvider):
    def __init__(self) -> None:
        self.revise_calls = 0

    @property
    def model(self) -> str:
        return "stub-model"

    def generate(
        self,
        rca_input: RCAInput,
        *,
        prompt_version: str,
        strict_retry: bool = False,
    ) -> RCAReport:
        return RCAReport.model_validate(
            report_payload(
                root_cause="A generic configuration failure in testing.",
                selected_cause=(
                    "The migration release path lacked a performance validation gate."
                ),
            )
        ).model_copy(
            update={
                "source_model": self.model,
                "prompt_version": prompt_version,
                "latency_seconds": 0.1,
            }
        )

    def create_structured(
        self,
        messages: list[dict[str, Any]],
        response_model: type[RCAReport],
        *,
        max_retries: int | None = None,
    ) -> RCAReport:
        self.revise_calls += 1
        return response_model.model_validate(
            report_payload(
                root_cause=(
                    "The migration release path lacked a performance validation gate."
                ),
                selected_cause=(
                    "The migration release path lacked a performance validation gate."
                ),
            )
        )


def test_agent_revises_method_consistency_findings() -> None:
    provider = StubProvider()
    settings = Settings(
        validation_enabled=False,
        max_revise_rounds=2,
        agent_timeout_seconds=60,
        memory_writeback_enabled=False,
    )
    agent = RCAAgent(settings=settings, provider=provider)

    report = agent.run(
        "Checkout requests time out after a database migration.",
        method="fishbone",
    )

    assert provider.revise_calls == 1
    assert report.method_detail is not None
    assert report.method_detail.fishbone is not None
    assert report.root_cause == report.method_detail.fishbone.selected_cause
    assert any("method_consistency" in note for note in report.validation_notes)


def test_agent_appends_completed_run_to_memory_workbook(tmp_path) -> None:
    provider = StubProvider()
    memory_path = tmp_path / "past_rca_memory.xlsx"
    settings = Settings(
        validation_enabled=False,
        max_revise_rounds=2,
        agent_timeout_seconds=60,
        memory_enabled=False,
        memory_writeback_enabled=True,
        memory_path=memory_path,
    )
    agent = RCAAgent(settings=settings, provider=provider)

    report = agent.run(
        "Checkout requests time out after a database migration.",
        context="Migration finished ten minutes before checkout alerts fired.",
        method="fishbone",
        severity="high",
        system_area="payments",
    )

    workbook = load_workbook(memory_path)
    sheet = workbook["Past RCA Memory"]
    headers = [cell.value for cell in sheet[1]]
    values = dict(zip(headers, [cell.value for cell in sheet[2]], strict=False))

    assert sheet.max_row == 2
    assert str(values["incident_id"]).startswith("AUTO-")
    assert values["problem_statement"] == "Checkout requests time out after a database migration."
    assert values["system_area"] == "payments"
    assert values["root_cause"] == report.root_cause
    assert values["confidence"] == report.confidence
    assert values["status"] == "generated"
    assert "method:fishbone" in values["tags"]
    assert any("added this RCA to past RCA memory" in note for note in report.validation_notes)
    assert agent.last_run_stats["memory_writeback"]["row_number"] == 2


def test_memory_evidence_respects_total_context_budget(monkeypatch, tmp_path) -> None:
    provider = CapturingStubProvider()
    settings = Settings(
        validation_enabled=False,
        max_revise_rounds=0,
        memory_enabled=True,
        memory_writeback_enabled=False,
        memory_path=tmp_path / "unused.xlsx",
        max_context_chars=120,
    )
    monkeypatch.setattr(
        "agent.orchestrator.search_past_rca_memory",
        lambda *args, **kwargs: MemorySearch(
            matches=[],
            evidence_pack="past evidence " * 100,
            retrieval_mode="lexical",
        ),
    )
    agent = RCAAgent(settings=settings, provider=provider)

    agent.run(
        "Checkout requests fail after a database deployment.",
        context="x" * 90,
    )

    assert provider.seen_inputs
    assert len(provider.seen_inputs[0].context or "") <= settings.max_context_chars
    assert any(
        "truncated past_rca_memory" in finding
        for finding in agent.last_run_stats["sanitizer_findings"]
    )
