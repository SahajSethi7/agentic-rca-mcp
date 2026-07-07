"""Semantic (embedding) retrieval tests using a deterministic fake embedder.

The fake maps tokens to a handful of concept dimensions so paraphrases land
close together and unrelated text lands near-orthogonal - no Ollama needed.
"""

from __future__ import annotations

import re
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import embeddings as embeddings_module
from embeddings import (
    COSINE_CEILING,
    COSINE_FLOOR,
    EmbeddingConfig,
    ensure_memory_embedding_index,
    normalize_cosine,
    semantic_contribution,
)
from memory import MEMORY_COLUMNS, search_past_rca_memory
from schemas import RCAInput

_CONCEPTS: tuple[frozenset[str], ...] = (
    frozenset({"auth", "authentication", "login", "sign", "signin", "sso", "token", "jwt", "issuer", "locked", "accounts"}),
    frozenset({"deployment", "deploy", "rollout", "release", "shipped"}),
    frozenset({"invoice", "batch", "scheduler", "overnight", "jobs"}),
    frozenset({"database", "migration", "timeout", "latency"}),
)


def _fake_embed_factory(calls: list[list[str]]):
    def fake_embed(texts: list[str], config: EmbeddingConfig) -> list[list[float]]:
        calls.append(list(texts))
        vectors: list[list[float]] = []
        for text in texts:
            tokens = set(re.findall(r"[a-z0-9]+", text.lower()))
            concept_hits = [float(len(tokens & concept)) for concept in _CONCEPTS]
            # Small shared baseline keeps unrelated cosines positive but far
            # below the COSINE_FLOOR clamp.
            vectors.append([*concept_hits, 0.2])
        return vectors

    return fake_embed


def _memory_workbook(path: Path) -> None:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Past RCA Memory"
    sheet.append(MEMORY_COLUMNS)
    sheet.append(
        [
            "AUTH-001",
            "2026-07-01",
            "identity",
            "login-api",
            "JWT issuer mismatch",
            "Authentication requests fail with JWT issuer errors right when the deployment finished.",
            "Token validation errors during login.",
            "Deployment removed the JWT issuer environment variable.",
            "Restore the JWT issuer env var.",
            "Add deployment validation for auth env vars.",
            "deployment diff; auth logs",
            "identity-platform",
            "jwt",
            "high",
            "approved",
        ]
    )
    sheet.append(
        [
            "BATCH-001",
            "2026-07-02",
            "finance",
            "invoice-worker",
            "scheduler disabled",
            "Invoice batches stopped overnight.",
            "No invoices generated after scheduler change.",
            "Scheduler config was disabled in production.",
            "Re-enable scheduler.",
            "Add scheduler config drift alert.",
            "scheduler audit; job logs",
            "finance-platform",
            "batch,scheduler",
            "medium",
            "approved",
        ]
    )
    wb.save(path)


def _config(tmp_path: Path, **overrides) -> EmbeddingConfig:
    defaults = dict(
        base_url="http://localhost:11434/v1",
        model="fake-embed",
        index_path=tmp_path / "cache" / "rca_memory_embeddings.sqlite",
        weight=0.40,
    )
    defaults.update(overrides)
    return EmbeddingConfig(**defaults)


# Paraphrase of AUTH-001 sharing (almost) no useful tokens with the record.
PARAPHRASED_QUERY = "Customers are locked out of their accounts and cannot sign in since the rollout."


def test_semantic_retrieval_recovers_paraphrase_lexical_misses(tmp_path, monkeypatch) -> None:
    memory_path = tmp_path / "past_rca_memory.xlsx"
    _memory_workbook(memory_path)
    calls: list[list[str]] = []
    monkeypatch.setattr(embeddings_module, "embed_texts", _fake_embed_factory(calls))

    rca_input = RCAInput(problem_statement=PARAPHRASED_QUERY)

    lexical_only = search_past_rca_memory(
        rca_input,
        memory_path,
        max_matches=3,
        min_score=0.20,
        graph_enabled=False,
        graph_path=tmp_path / "cache" / "graph.sqlite",
    )
    assert not [m for m in lexical_only.matches if m.incident_id == "AUTH-001"]

    semantic = search_past_rca_memory(
        rca_input,
        memory_path,
        max_matches=3,
        min_score=0.20,
        graph_enabled=False,
        graph_path=tmp_path / "cache" / "graph.sqlite",
        embedding_config=_config(tmp_path),
    )
    assert semantic.matches
    top = semantic.matches[0]
    assert top.incident_id == "AUTH-001"
    assert top.retrieval_mode in {"semantic", "hybrid"}
    assert "semantic similarity" in top.match_reason
    # Capped: embeddings alone can never dominate the hybrid score.
    assert top.similarity_score <= 0.40 + 1e-9


def test_semantic_retrieval_fails_soft_when_endpoint_down(tmp_path) -> None:
    memory_path = tmp_path / "past_rca_memory.xlsx"
    _memory_workbook(memory_path)

    result = search_past_rca_memory(
        RCAInput(problem_statement="Invoice batch scheduler disabled in production."),
        memory_path,
        max_matches=2,
        min_score=0.20,
        graph_enabled=False,
        graph_path=tmp_path / "cache" / "graph.sqlite",
        # Unroutable local port: embed call raises, retrieval must fall back.
        embedding_config=_config(tmp_path, base_url="http://127.0.0.1:9/v1", timeout_seconds=1.0),
    )

    assert result.matches[0].incident_id == "BATCH-001"
    assert result.matches[0].retrieval_mode == "lexical"
    assert result.warning and "embeddings unavailable" in result.warning.lower()


def test_embedding_index_reports_fresh_after_rebuild(tmp_path, monkeypatch) -> None:
    memory_path = tmp_path / "past_rca_memory.xlsx"
    _memory_workbook(memory_path)
    calls: list[list[str]] = []
    monkeypatch.setattr(embeddings_module, "embed_texts", _fake_embed_factory(calls))

    records = [
        {
            "incident_id": "AUTH-001",
            "problem_statement": "Authentication requests fail after deployment.",
            "symptoms": "Token validation errors during login.",
            "error_signature": "JWT issuer mismatch",
        },
        {
            "incident_id": "BATCH-001",
            "problem_statement": "Invoice batches stopped overnight.",
            "symptoms": "No invoices generated after scheduler change.",
            "error_signature": "scheduler disabled",
        },
    ]

    status = ensure_memory_embedding_index(memory_path, _config(tmp_path), records)

    assert status["fresh"] is True
    assert status["embedded_count"] == 2
    assert status["vector_count"] == 2


def test_embedding_index_is_incremental_on_writeback_append(tmp_path, monkeypatch) -> None:
    memory_path = tmp_path / "past_rca_memory.xlsx"
    _memory_workbook(memory_path)
    calls: list[list[str]] = []
    monkeypatch.setattr(embeddings_module, "embed_texts", _fake_embed_factory(calls))
    config = _config(tmp_path)
    rca_input = RCAInput(problem_statement=PARAPHRASED_QUERY)

    def run_search() -> None:
        search_past_rca_memory(
            rca_input,
            memory_path,
            max_matches=3,
            min_score=0.20,
            graph_enabled=False,
            graph_path=tmp_path / "cache" / "graph.sqlite",
            embedding_config=config,
        )

    run_search()
    embedded_texts = [text for batch in calls for text in batch]
    assert len(embedded_texts) == 3  # two memory rows + the query

    calls.clear()
    run_search()  # index fresh: only the query is embedded
    assert [len(batch) for batch in calls] == [1]

    # Simulate a memory write-back append.
    wb = load_workbook(memory_path)
    sheet = wb["Past RCA Memory"]
    sheet.append(
        [
            "AUTO-001",
            "2026-07-07",
            "identity",
            "sso-gateway",
            "SAML assertion expired",
            "SSO sign-in loops back to the login page.",
            "SAML assertion rejected as expired.",
            "Clock drift on the gateway host.",
            "Resync NTP.",
            "Alert on clock drift.",
            "gateway logs",
            "identity-platform",
            "sso",
            "medium",
            "generated",
        ]
    )
    wb.save(memory_path)

    calls.clear()
    run_search()
    embedded_texts = [text for batch in calls for text in batch]
    # Only the new row plus the query - old rows are reused via row hashes.
    assert len(embedded_texts) == 2
    assert any("SSO sign-in loops" in text for text in embedded_texts)


def test_cosine_is_clamped_thresholded_and_gated() -> None:
    config = EmbeddingConfig(
        base_url="http://localhost:11434/v1",
        model="fake-embed",
        index_path=Path("unused.sqlite"),
        weight=0.40,
        lexical_gate=0.35,
    )
    # Threshold: at/below the floor there is no signal at all.
    assert normalize_cosine(COSINE_FLOOR) == 0.0
    assert normalize_cosine(0.30) == 0.0
    # Clamp: values above the ceiling saturate at 1.
    assert normalize_cosine(COSINE_CEILING) == pytest.approx(1.0)
    assert normalize_cosine(0.99) == pytest.approx(1.0)
    mid = (COSINE_FLOOR + COSINE_CEILING) / 2
    assert normalize_cosine(mid) == pytest.approx(0.5)

    # Full weight only when lexical evidence is weak.
    assert semantic_contribution(0.99, 0.0, config) == pytest.approx(0.40)
    # Strong lexical evidence: semantic only tops up at quarter strength.
    assert semantic_contribution(0.99, 0.60, config) == pytest.approx(0.10)
    # Cap holds everywhere.
    assert semantic_contribution(2.0, 0.0, config) <= 0.40
