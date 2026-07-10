from __future__ import annotations

from conftest import clean_report_payload
from openpyxl import load_workbook

from memory import append_rca_to_memory, build_memory_matches_workbook
from schemas import KnownIssueMatch, RCAInput, RCAReport


def test_memory_writeback_neutralizes_spreadsheet_formulas(tmp_path) -> None:
    path = tmp_path / "memory.xlsx"
    rca_input = RCAInput(
        problem_statement="=HYPERLINK(\"https://example.invalid\",\"incident payload\")",
        context="@SUM(1+1) malicious spreadsheet payload",
        system_area="+cmd|' /C calc'!A0",
    )
    report = RCAReport.model_validate(clean_report_payload()).model_copy(
        update={"root_cause": "-Spreadsheet formula shaped root cause from untrusted model output"}
    )

    append_rca_to_memory(rca_input, report, path)

    sheet = load_workbook(path, data_only=False)["Past RCA Memory"]
    headers = [cell.value for cell in sheet[1]]
    values = dict(zip(headers, [cell.value for cell in sheet[2]], strict=False))
    for key in ("problem_statement", "symptoms", "system_area", "root_cause"):
        assert str(values[key]).startswith("'")
        assert sheet.cell(2, headers.index(key) + 1).data_type != "f"


def test_matching_workbook_neutralizes_all_untrusted_text(tmp_path) -> None:
    path = tmp_path / "matches.xlsx"
    match = KnownIssueMatch(
        incident_id="=INCIDENT()",
        problem_statement="+malicious prior problem",
        root_cause="@malicious prior root cause",
        similarity_score=0.9,
        match_reason="-malicious match reason",
    )

    build_memory_matches_workbook(
        [match],
        path,
        current_problem="=CURRENT_PROBLEM()",
        min_score=0.5,
    )

    workbook = load_workbook(path, data_only=False)
    assert str(workbook["Summary"]["B3"].value).startswith("'")
    matches = workbook["Matching past RCAs"]
    for coordinate in ("B2", "G2", "I2", "S2"):
        assert str(matches[coordinate].value).startswith("'")
        assert matches[coordinate].data_type != "f"
