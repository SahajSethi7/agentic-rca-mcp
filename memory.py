"""Read-only past RCA memory retrieval for demo grounding.

The first version intentionally stays local and deterministic: it reads an Excel
workbook, ranks similar incidents with lexical matching, and returns a compact
evidence pack. If LangGraph/LangChain are installed, the same steps are wrapped
in a tiny graph so the demo path can show the intended orchestration shape.
"""

from __future__ import annotations

import math
import re
import sqlite3
import threading
from dataclasses import dataclass
from datetime import datetime, timezone
from hashlib import sha256
from pathlib import Path
from typing import Any, TypedDict

from schemas import KnownIssueMatch, RCAInput, RCAReport

MEMORY_COLUMNS = [
    "incident_id",
    "date",
    "system_area",
    "service_name",
    "error_signature",
    "problem_statement",
    "symptoms",
    "root_cause",
    "immediate_fix",
    "long_term_fix",
    "evidence_checked",
    "owner_team",
    "tags",
    "confidence",
    "status",
]

TOKEN_RE = re.compile(r"[a-z0-9]+")
STOPWORDS = {
    "a",
    "an",
    "and",
    "are",
    "as",
    "at",
    "be",
    "by",
    "for",
    "from",
    "in",
    "is",
    "it",
    "of",
    "on",
    "or",
    "that",
    "the",
    "this",
    "to",
    "was",
    "were",
    "with",
    "after",
    "before",
    "during",
    "due",
    "issue",
    "problem",
    "failure",
}

GENERIC_ROOT_CAUSES = {
    "configuration issue",
    "configuration problem",
    "configuration failure",
    "monitoring issue",
    "monitoring failure",
    "process issue",
    "process gap",
    "validation issue",
    "validation gap",
    "testing issue",
    "operational issue",
}

_WRITEBACK_LOCK = threading.Lock()


@dataclass(frozen=True)
class MemorySearch:
    matches: list[KnownIssueMatch]
    evidence_pack: str | None
    retrieval_mode: str
    context_match_count: int = 0
    warning: str | None = None


@dataclass(frozen=True)
class MemoryWriteBack:
    incident_id: str
    row_number: int
    memory_path: Path


class MemoryState(TypedDict, total=False):
    rca_input: RCAInput
    memory_path: Path
    max_matches: int
    min_score: float
    records: list[dict[str, Any]]
    matches: list[KnownIssueMatch]
    context_matches: list[KnownIssueMatch]
    evidence_pack: str | None
    graph_enabled: bool
    graph_path: Path
    graph_warning: str | None


def _tokens(text: str | None) -> set[str]:
    if not text:
        return set()
    return {token for token in TOKEN_RE.findall(text.lower()) if len(token) > 2 and token not in STOPWORDS}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def _date_text(value: Any) -> str | None:
    text = _text(value)
    if not text:
        return None
    if " " in text:
        return text.split(" ", 1)[0]
    return text


def _confidence_value(value: str) -> str | None:
    normalized = value.strip().lower()
    if normalized in {"low", "medium", "high"}:
        return normalized
    return None


def _query_text(rca_input: RCAInput) -> str:
    return " ".join(
        part
        for part in [
            rca_input.problem_statement,
            rca_input.context or "",
            rca_input.system_area or "",
            rca_input.severity or "",
        ]
        if part
    )


GRAPH_FIELDS: tuple[tuple[str, str, str, float], ...] = (
    ("system_area", "system_area", "has_system_area", 0.16),
    ("service_name", "service", "affects_service", 0.16),
    ("error_signature", "error_signature", "has_signature", 0.18),
    ("root_cause", "root_cause", "caused_by", 0.10),
    ("immediate_fix", "fix", "fixed_by", 0.09),
    ("long_term_fix", "fix", "prevented_by", 0.08),
    ("evidence_checked", "evidence", "checked_by_evidence", 0.07),
    ("owner_team", "owner_team", "owned_by", 0.05),
    ("confidence", "confidence", "has_confidence", 0.03),
    ("status", "status", "has_status", 0.03),
)


def _normalize_key(value: str) -> str:
    return " ".join(TOKEN_RE.findall(value.lower()))


def _split_tags(value: str | None) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in re.split(r"[,;|]", value) if part.strip()]


def _field_values(record: dict[str, Any], field: str) -> list[str]:
    if field == "tags":
        return _split_tags(record.get("tags"))
    value = _text(record.get(field))
    return [value] if value else []


def _memory_fingerprint(memory_path: Path) -> dict[str, str]:
    stat = memory_path.stat()
    return {
        "source_path": str(memory_path.resolve()),
        "source_mtime_ns": str(stat.st_mtime_ns),
        "source_size": str(stat.st_size),
    }


def _graph_meta(conn: sqlite3.Connection) -> dict[str, str]:
    try:
        return {
            str(key): str(value)
            for key, value in conn.execute("SELECT key, value FROM meta").fetchall()
        }
    except sqlite3.Error:
        return {}


def _init_graph_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS meta (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS nodes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kind TEXT NOT NULL,
            key TEXT NOT NULL,
            label TEXT NOT NULL,
            UNIQUE(kind, key)
        );
        CREATE TABLE IF NOT EXISTS edges (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_id INTEGER NOT NULL,
            target_id INTEGER NOT NULL,
            kind TEXT NOT NULL,
            weight REAL NOT NULL,
            incident_id TEXT NOT NULL,
            FOREIGN KEY(source_id) REFERENCES nodes(id),
            FOREIGN KEY(target_id) REFERENCES nodes(id)
        );
        CREATE INDEX IF NOT EXISTS idx_nodes_kind_key ON nodes(kind, key);
        CREATE INDEX IF NOT EXISTS idx_edges_incident ON edges(incident_id);
        """
    )


def _node_id(conn: sqlite3.Connection, *, kind: str, label: str) -> int:
    key = _normalize_key(label)
    conn.execute(
        "INSERT OR IGNORE INTO nodes(kind, key, label) VALUES (?, ?, ?)",
        (kind, key, label),
    )
    row = conn.execute(
        "SELECT id FROM nodes WHERE kind = ? AND key = ?",
        (kind, key),
    ).fetchone()
    if row is None:
        raise RuntimeError(f"Failed to create graph node for {kind}:{label}")
    return int(row[0])


def _rebuild_graph_index(conn: sqlite3.Connection, memory_path: Path, records: list[dict[str, Any]]) -> None:
    conn.executescript("DELETE FROM edges; DELETE FROM nodes; DELETE FROM meta;")
    for record in records:
        incident_id = _text(record.get("incident_id"))
        if not incident_id:
            continue
        incident_node = _node_id(conn, kind="incident", label=incident_id)
        for field, kind, edge_kind, weight in GRAPH_FIELDS:
            for value in _field_values(record, field):
                target = _node_id(conn, kind=kind, label=value)
                conn.execute(
                    """
                    INSERT INTO edges(source_id, target_id, kind, weight, incident_id)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (incident_node, target, edge_kind, weight, incident_id),
                )
        for tag in _split_tags(record.get("tags")):
            target = _node_id(conn, kind="tag", label=tag)
            conn.execute(
                "INSERT INTO edges(source_id, target_id, kind, weight, incident_id) VALUES (?, ?, ?, ?, ?)",
                (incident_node, target, "tagged_with", 0.10, incident_id),
            )

    for key, value in _memory_fingerprint(memory_path).items():
        conn.execute("INSERT INTO meta(key, value) VALUES (?, ?)", (key, value))
    conn.execute("INSERT INTO meta(key, value) VALUES (?, ?)", ("record_count", str(len(records))))


def ensure_memory_graph_index(
    memory_path: Path,
    graph_path: Path,
    records: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Create or refresh the derived SQLite graph cache for the Excel memory."""
    if records is None:
        records = _load_records(memory_path)
    graph_path.parent.mkdir(parents=True, exist_ok=True)
    expected = _memory_fingerprint(memory_path)
    conn = sqlite3.connect(graph_path)
    try:
        _init_graph_schema(conn)
        if _graph_meta(conn) != {**expected, "record_count": str(len(records))}:
            _rebuild_graph_index(conn, memory_path, records)
            rebuilt = True
        else:
            rebuilt = False
        node_count = conn.execute("SELECT COUNT(*) FROM nodes").fetchone()[0]
        edge_count = conn.execute("SELECT COUNT(*) FROM edges").fetchone()[0]
        conn.commit()
    finally:
        conn.close()
    return {
        "enabled": True,
        "path": str(graph_path),
        "exists": graph_path.exists(),
        "fresh": True,
        "rebuilt": rebuilt,
        "node_count": node_count,
        "edge_count": edge_count,
        "record_count": len(records),
        "warning": None,
    }


def get_memory_graph_status(memory_path: Path, graph_path: Path, *, enabled: bool) -> dict[str, Any]:
    status: dict[str, Any] = {
        "enabled": enabled,
        "path": str(graph_path),
        "exists": graph_path.exists(),
        "fresh": False,
        "node_count": None,
        "edge_count": None,
        "record_count": None,
        "warning": None,
    }
    if not enabled:
        return status
    if not memory_path.exists():
        status["warning"] = f"RCA memory file not found: {memory_path}"
        return status
    if not graph_path.exists():
        status["warning"] = "Graph index has not been built yet."
        return status
    try:
        expected = _memory_fingerprint(memory_path)
        conn = sqlite3.connect(graph_path)
        try:
            _init_graph_schema(conn)
            meta = _graph_meta(conn)
            status["fresh"] = all(meta.get(key) == value for key, value in expected.items())
            status["record_count"] = int(meta.get("record_count", "0") or 0)
            status["node_count"] = conn.execute("SELECT COUNT(*) FROM nodes").fetchone()[0]
            status["edge_count"] = conn.execute("SELECT COUNT(*) FROM edges").fetchone()[0]
            conn.commit()
        finally:
            conn.close()
        if not status["fresh"]:
            status["warning"] = "Graph index is stale and will be rebuilt on the next memory search."
    except Exception as exc:  # noqa: BLE001 - status should explain rather than crash.
        status["warning"] = f"{type(exc).__name__}: {exc}"
    return status


def _load_records(path: Path) -> list[dict[str, Any]]:
    import pandas as pd

    frame = pd.read_excel(path, sheet_name="Past RCA Memory")
    missing = [column for column in MEMORY_COLUMNS if column not in frame.columns]
    if missing:
        raise ValueError(f"RCA memory sheet is missing columns: {', '.join(missing)}")
    records: list[dict[str, Any]] = []
    for raw in frame[MEMORY_COLUMNS].to_dict(orient="records"):
        record = {key: _text(value) for key, value in raw.items()}
        record["date"] = _date_text(raw.get("date"))
        records.append(record)
    return records


def _score_record(query_tokens: set[str], rca_input: RCAInput, record: dict[str, Any]) -> tuple[float, str]:
    """Score a memory row with a balanced lexical/signal match.

    The score is intentionally not model confidence. It is only a retrieval
    relevance score used to surface past RCA memory.
    """
    problem_tokens = _tokens(record.get("problem_statement"))
    symptom_tokens = _tokens(record.get("symptoms"))
    signature_tokens = _tokens(record.get("error_signature"))
    root_tokens = _tokens(record.get("root_cause"))
    service_tokens = _tokens(record.get("service_name"))
    system_tokens = _tokens(record.get("system_area"))
    tag_tokens = _tokens(record.get("tags"))
    fix_tokens = _tokens(
        " ".join([record.get("immediate_fix", ""), record.get("long_term_fix", "")])
    )
    record_tokens = (
        problem_tokens
        | symptom_tokens
        | signature_tokens
        | root_tokens
        | service_tokens
        | system_tokens
        | tag_tokens
        | fix_tokens
    )
    if not query_tokens or not record_tokens:
        return 0.0, "No useful token overlap."

    overlap_tokens = query_tokens & record_tokens
    recall = len(overlap_tokens) / max(len(query_tokens), 1)
    jaccard = len(overlap_tokens) / max(len(query_tokens | record_tokens), 1)
    problem_overlap = len(query_tokens & problem_tokens) / max(len(query_tokens), 1)
    symptom_overlap = len(query_tokens & symptom_tokens) / max(len(query_tokens), 1)
    signature_overlap = len(query_tokens & signature_tokens) / max(len(query_tokens), 1)
    signal_overlap = len(query_tokens & (service_tokens | system_tokens | tag_tokens)) / max(len(query_tokens), 1)

    score = (
        0.38 * recall
        + 0.22 * jaccard
        + 0.18 * problem_overlap
        + 0.12 * symptom_overlap
        + 0.06 * signature_overlap
        + 0.04 * signal_overlap
    )

    reasons: list[str] = []
    if overlap_tokens:
        shared = sorted(overlap_tokens)
        reasons.append("shared terms: " + ", ".join(shared[:8]))

    input_system_tokens = _tokens(rca_input.system_area)
    if input_system_tokens and input_system_tokens & system_tokens:
        score += 0.12
        reasons.append(f"same system area: {record.get('system_area')}")

    if service_tokens & query_tokens:
        score += 0.08
        reasons.append(f"service hint: {record.get('service_name')}")

    if tag_tokens & query_tokens:
        score += 0.06
        reasons.append("tag overlap")

    if signature_tokens & query_tokens:
        score += 0.08
        reasons.append("error signature overlap")

    if root_tokens & query_tokens:
        score += 0.04
        reasons.append("root-cause term overlap")

    return min(score, 1.0), "; ".join(reasons) if reasons else "Weak lexical similarity."


def _graph_score_record(query_tokens: set[str], record: dict[str, Any]) -> tuple[float, list[str], list[str]]:
    if not query_tokens:
        return 0.0, [], []
    score = 0.0
    reasons: list[str] = []
    paths: list[str] = []
    incident_id = record.get("incident_id", "unknown")

    for field, kind, edge_kind, weight in (*GRAPH_FIELDS, ("tags", "tag", "tagged_with", 0.10)):
        for value in _field_values(record, field):
            value_tokens = _tokens(value)
            overlap = query_tokens & value_tokens
            if not overlap:
                continue
            signal = min(weight, weight * (0.5 + len(overlap) / max(len(value_tokens), 1)))
            score += signal
            label = value if len(value) <= 84 else value[:81].rstrip() + "..."
            reasons.append(f"{kind} graph signal: {label}")
            if len(paths) < 3:
                paths.append(f"current incident -> {kind}:{label} -> {edge_kind} -> {incident_id}")

    return min(score, 0.42), reasons, paths


def _match_from_record(
    *,
    score: float,
    reason: str,
    record: dict[str, Any],
    retrieval_mode: str,
    graph_path: list[str],
) -> KnownIssueMatch:
    return KnownIssueMatch(
        incident_id=record.get("incident_id", ""),
        date=record.get("date") or None,
        system_area=record.get("system_area") or None,
        service_name=record.get("service_name") or None,
        error_signature=record.get("error_signature") or None,
        problem_statement=record.get("problem_statement", ""),
        symptoms=record.get("symptoms") or None,
        root_cause=record.get("root_cause", ""),
        immediate_fix=record.get("immediate_fix") or None,
        long_term_fix=record.get("long_term_fix") or None,
        evidence_checked=record.get("evidence_checked") or None,
        owner_team=record.get("owner_team") or None,
        tags=record.get("tags") or None,
        confidence=_confidence_value(record.get("confidence", "")),
        status=record.get("status") or None,
        similarity_score=round(score, 3),
        match_reason=reason,
        retrieval_mode=retrieval_mode,  # type: ignore[arg-type]
        graph_path=graph_path,
    )


def _rank_records(state: MemoryState) -> MemoryState:
    rca_input = state["rca_input"]
    query_tokens = _tokens(_query_text(rca_input))
    graph_warning: str | None = None
    graph_enabled = bool(state.get("graph_enabled"))
    if graph_enabled:
        try:
            ensure_memory_graph_index(
                state["memory_path"],
                state["graph_path"],
                state.get("records", []),
            )
        except Exception as exc:  # noqa: BLE001 - graph retrieval must fail soft.
            graph_warning = f"RCA memory graph unavailable; used lexical retrieval ({type(exc).__name__})."
            graph_enabled = False

    scored: list[tuple[float, str, str, list[str], dict[str, Any]]] = []
    for record in state.get("records", []):
        lexical_score, lexical_reason = _score_record(query_tokens, rca_input, record)
        graph_score, graph_reasons, graph_path = (
            _graph_score_record(query_tokens, record) if graph_enabled else (0.0, [], [])
        )
        score = min(1.0, lexical_score + graph_score)
        if graph_score and lexical_score:
            retrieval_mode = "hybrid"
        elif graph_score:
            retrieval_mode = "graph"
        else:
            retrieval_mode = "lexical"
        reasons = [lexical_reason] if lexical_score else []
        reasons.extend(graph_reasons[:3])
        reason = "; ".join(reasons) if reasons else "Weak lexical similarity."
        if score >= state["min_score"]:
            scored.append((score, reason, retrieval_mode, graph_path, record))
    scored.sort(key=lambda item: item[0], reverse=True)

    matches: list[KnownIssueMatch] = []
    for score, reason, retrieval_mode, graph_path, record in scored:
        matches.append(
            _match_from_record(
                score=score,
                reason=reason,
                record=record,
                retrieval_mode=retrieval_mode,
                graph_path=graph_path,
            )
        )
    context_limit = max(0, state["max_matches"])
    return {
        **state,
        "matches": matches,
        "context_matches": matches[:context_limit],
        "graph_warning": graph_warning,
        "graph_enabled": graph_enabled,
    }


def _build_evidence_pack(state: MemoryState) -> MemoryState:
    matches = state.get("context_matches", [])
    if not matches:
        return {**state, "evidence_pack": None}

    lines = [
        "PAST RCA MEMORY MATCHES (read-only local Excel and graph retrieval)",
        "Use these records as supporting evidence only when the current symptoms are similar.",
        "Do not copy a past fix blindly; reason from the current incident.",
        "",
    ]
    for index, match in enumerate(matches, start=1):
        lines.extend(
            [
                f"Match {index}: {match.incident_id} (score {match.similarity_score:.2f})",
                f"- retrieval: {match.retrieval_mode}",
                f"- service/signature: {match.service_name or 'unknown'} / {match.error_signature or 'not recorded'}",
                f"- past root cause: {match.root_cause}",
                f"- past fix signal: {match.immediate_fix or 'not recorded'}",
                f"- graph path: {' | '.join(match.graph_path) if match.graph_path else 'not recorded'}",
                "",
            ]
        )
    return {**state, "evidence_pack": "\n".join(lines).strip()}


def _direct_memory_search(
    rca_input: RCAInput,
    memory_path: Path,
    *,
    max_matches: int,
    min_score: float,
    graph_enabled: bool,
    graph_path: Path,
) -> MemorySearch:
    records = _load_records(memory_path)
    state: MemoryState = {
        "rca_input": rca_input,
        "memory_path": memory_path,
        "max_matches": max_matches,
        "min_score": min_score,
        "records": records,
        "graph_enabled": graph_enabled,
        "graph_path": graph_path,
    }
    state = _rank_records(state)
    state = _build_evidence_pack(state)
    return MemorySearch(
        matches=state.get("matches", []),
        evidence_pack=state.get("evidence_pack"),
        retrieval_mode="graph" if state.get("graph_enabled") else "deterministic",
        context_match_count=len(state.get("context_matches", [])),
        warning=state.get("graph_warning"),
    )


def _langgraph_memory_search(
    rca_input: RCAInput,
    memory_path: Path,
    *,
    max_matches: int,
    min_score: float,
    graph_enabled: bool,
    graph_path: Path,
) -> MemorySearch:
    from langchain_core.documents import Document
    from langgraph.graph import END, START, StateGraph

    def load_node(state: MemoryState) -> MemoryState:
        records = _load_records(state["memory_path"])
        # Keep a LangChain Document representation in the graph boundary so the
        # orchestration mirrors the intended future tool/retriever shape.
        _ = [
            Document(
                page_content=" ".join(
                    [
                        record.get("problem_statement", ""),
                        record.get("symptoms", ""),
                        record.get("root_cause", ""),
                    ]
                ),
                metadata={
                    "incident_id": record.get("incident_id"),
                    "service_name": record.get("service_name"),
                    "system_area": record.get("system_area"),
                },
            )
            for record in records
        ]
        return {**state, "records": records}

    graph = StateGraph(MemoryState)
    graph.add_node("load_excel_memory", load_node)
    graph.add_node("rank_similar_incidents", _rank_records)
    graph.add_node("build_evidence_pack", _build_evidence_pack)
    graph.add_edge(START, "load_excel_memory")
    graph.add_edge("load_excel_memory", "rank_similar_incidents")
    graph.add_edge("rank_similar_incidents", "build_evidence_pack")
    graph.add_edge("build_evidence_pack", END)
    compiled = graph.compile()
    state = compiled.invoke(
        {
            "rca_input": rca_input,
            "memory_path": memory_path,
            "max_matches": max_matches,
            "min_score": min_score,
            "graph_enabled": graph_enabled,
            "graph_path": graph_path,
        }
    )
    return MemorySearch(
        matches=state.get("matches", []),
        evidence_pack=state.get("evidence_pack"),
        retrieval_mode="langgraph-graph" if state.get("graph_enabled") else "langgraph",
        context_match_count=len(state.get("context_matches", [])),
        warning=state.get("graph_warning"),
    )


def search_past_rca_memory(
    rca_input: RCAInput,
    memory_path: Path,
    *,
    max_matches: int = 10,
    min_score: float = 0.50,
    graph_enabled: bool = True,
    graph_path: Path | None = None,
) -> MemorySearch:
    """Return all threshold-passing RCA records and a compact prompt evidence pack."""
    graph_path = graph_path or memory_path.with_suffix(".graph.sqlite")
    if not memory_path.exists():
        return MemorySearch(
            matches=[],
            evidence_pack=None,
            retrieval_mode="disabled",
            context_match_count=0,
            warning=f"RCA memory file not found: {memory_path}",
        )

    try:
        return _langgraph_memory_search(
            rca_input,
            memory_path,
            max_matches=max_matches,
            min_score=min_score,
            graph_enabled=graph_enabled,
            graph_path=graph_path,
        )
    except ModuleNotFoundError as exc:
        try:
            result = _direct_memory_search(
                rca_input,
                memory_path,
                max_matches=max_matches,
                min_score=min_score,
                graph_enabled=graph_enabled,
                graph_path=graph_path,
            )
            return MemorySearch(
                matches=result.matches,
                evidence_pack=result.evidence_pack,
                retrieval_mode=result.retrieval_mode,
                context_match_count=result.context_match_count,
                warning=result.warning
                or f"LangGraph memory path unavailable; used direct fallback ({type(exc).__name__}).",
            )
        except Exception as fallback_exc:
            return MemorySearch(
                matches=[],
                evidence_pack=None,
                retrieval_mode="error",
                context_match_count=0,
                warning=f"RCA memory retrieval failed: {type(fallback_exc).__name__}",
            )
    except Exception as exc:
        try:
            result = _direct_memory_search(
                rca_input,
                memory_path,
                max_matches=max_matches,
                min_score=min_score,
                graph_enabled=graph_enabled,
                graph_path=graph_path,
            )
            return MemorySearch(
                matches=result.matches,
                evidence_pack=result.evidence_pack,
                retrieval_mode=result.retrieval_mode,
                context_match_count=result.context_match_count,
                warning=result.warning
                or f"LangGraph memory path unavailable; used direct fallback ({type(exc).__name__}).",
            )
        except Exception as fallback_exc:
            return MemorySearch(
                matches=[],
                evidence_pack=None,
                retrieval_mode="error",
                context_match_count=0,
                warning=f"RCA memory retrieval failed: {type(fallback_exc).__name__}",
            )


def get_past_rca_memory_count(memory_path: Path) -> int:
    """Return the number of usable RCA memory records in the workbook."""
    return sum(1 for record in _load_records(memory_path) if record.get("incident_id"))


def _clip_excel_text(value: str | None, limit: int = 8000) -> str:
    text = _text(value)
    if len(text) <= limit:
        return text
    return text[: limit - 14].rstrip() + " ...[truncated]"


def _join_short(items: list[str], *, limit: int = 8000) -> str:
    return _clip_excel_text("; ".join(item.strip() for item in items if item.strip()), limit)


def _generated_incident_id(rca_input: RCAInput, report: RCAReport, timestamp: datetime) -> str:
    digest = sha256(
        "|".join(
            [
                rca_input.problem_statement,
                report.root_cause,
                report.method or rca_input.method,
                timestamp.isoformat(timespec="seconds"),
            ]
        ).encode("utf-8")
    ).hexdigest()[:8].upper()
    return f"AUTO-{timestamp.strftime('%Y%m%d-%H%M%S')}-{digest}"


def _build_memory_writeback_record(
    rca_input: RCAInput,
    report: RCAReport,
    *,
    timestamp: datetime,
) -> dict[str, str]:
    recommendations = [item for item in report.recommendations if item.strip()]
    tags = [
        f"method:{report.method or rca_input.method}",
        "source:auto-rca-run",
    ]
    if rca_input.severity:
        tags.append(f"severity:{rca_input.severity}")
    if rca_input.system_area:
        tags.append(f"system_area:{rca_input.system_area}")
    if report.source_model:
        tags.append(f"model:{report.source_model}")

    evidence_parts = []
    if report.evidence_needed:
        evidence_parts.append("Evidence needed: " + _join_short(report.evidence_needed, limit=4000))
    if report.validation_notes:
        evidence_parts.append("Validation notes: " + _join_short(report.validation_notes, limit=4000))

    return {
        "incident_id": _generated_incident_id(rca_input, report, timestamp),
        "date": timestamp.date().isoformat(),
        "system_area": _clip_excel_text(rca_input.system_area),
        "service_name": "",
        "error_signature": "",
        "problem_statement": _clip_excel_text(rca_input.problem_statement),
        "symptoms": _clip_excel_text(rca_input.context or ""),
        "root_cause": _clip_excel_text(report.root_cause),
        "immediate_fix": _clip_excel_text(recommendations[0] if recommendations else ""),
        "long_term_fix": _join_short(recommendations[1:]),
        "evidence_checked": _clip_excel_text(" | ".join(evidence_parts)),
        "owner_team": "",
        "tags": _join_short(tags, limit=2000),
        "confidence": report.confidence,
        "status": "generated",
    }


def append_rca_to_memory(
    rca_input: RCAInput,
    report: RCAReport,
    memory_path: Path,
) -> MemoryWriteBack:
    """Append a completed RCA run to the past-RCA Excel memory workbook.

    This is intentionally separate from retrieval so write-back can be enabled
    only when the operator wants generated reports to become future memory.
    """
    from openpyxl import Workbook, load_workbook

    timestamp = datetime.now(timezone.utc)
    record = _build_memory_writeback_record(rca_input, report, timestamp=timestamp)

    with _WRITEBACK_LOCK:
        memory_path.parent.mkdir(parents=True, exist_ok=True)
        if memory_path.exists():
            workbook = load_workbook(memory_path)
            if "Past RCA Memory" in workbook.sheetnames:
                sheet = workbook["Past RCA Memory"]
            else:
                sheet = workbook.create_sheet("Past RCA Memory")
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Past RCA Memory"

        existing_headers = [
            _text(sheet.cell(row=1, column=column).value)
            for column in range(1, sheet.max_column + 1)
        ]
        if not any(existing_headers):
            for column, header in enumerate(MEMORY_COLUMNS, start=1):
                sheet.cell(row=1, column=column, value=header)
            headers = MEMORY_COLUMNS
        else:
            headers = existing_headers
            for header in MEMORY_COLUMNS:
                if header not in headers:
                    headers.append(header)
                    sheet.cell(row=1, column=len(headers), value=header)

        row_number = sheet.max_row + 1
        for column, header in enumerate(headers, start=1):
            sheet.cell(row=row_number, column=column, value=record.get(header, ""))

        workbook.save(memory_path)

    return MemoryWriteBack(
        incident_id=record["incident_id"],
        row_number=row_number,
        memory_path=memory_path,
    )


def build_memory_matches_workbook(
    matches: list[KnownIssueMatch],
    output_path: str | Path,
    *,
    current_problem: str,
    min_score: float,
) -> Path:
    """Write a downloadable workbook containing only the matched past RCAs."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    matches_sheet = wb.create_sheet("Matching past RCAs")

    title_fill = PatternFill("solid", fgColor="0F766E")
    header_fill = PatternFill("solid", fgColor="D1FAE5")
    pale_fill = PatternFill("solid", fgColor="F8FAFC")
    white_font = Font(color="FFFFFF", bold=True)
    header_font = Font(color="065F46", bold=True)
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    summary["A1"] = "RCA Assistant - Matching Past RCAs"
    summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = title_fill
    summary["A3"] = "Current problem statement"
    summary["B3"] = current_problem
    summary["A4"] = "Match threshold"
    summary["B4"] = f"{round(min_score * 100)}%"
    summary["A5"] = "Matching records"
    summary["B5"] = len(matches)
    summary["A6"] = "Important note"
    summary["B6"] = "Match percentage is retrieval similarity, not RCA verdict confidence."
    for row in range(3, 7):
        summary[f"A{row}"].font = bold
        summary[f"B{row}"].alignment = wrap
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 92

    headers = [
        "Match %",
        "Incident ID",
        "Date",
        "System Area",
        "Service",
        "Error Signature",
        "Problem Statement",
        "Symptoms",
        "Known Root Cause",
        "Immediate Fix",
        "Long Term Fix",
        "Evidence Checked",
        "Owner Team",
        "Tags",
        "Confidence",
        "Status",
        "Retrieval Mode",
        "Graph Path",
        "Match Reason",
    ]
    matches_sheet.append(headers)
    for cell in matches_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    if matches:
        for match in matches:
            matches_sheet.append(
                [
                    round(match.similarity_score * 100),
                    match.incident_id,
                    match.date or "",
                    match.system_area or "",
                    match.service_name or "",
                    match.error_signature or "",
                    match.problem_statement,
                    match.symptoms or "",
                    match.root_cause,
                    match.immediate_fix or "",
                    match.long_term_fix or "",
                    match.evidence_checked or "",
                    match.owner_team or "",
                    match.tags or "",
                    match.confidence or "",
                    match.status or "",
                    match.retrieval_mode,
                    " | ".join(match.graph_path),
                    match.match_reason,
                ]
            )
    else:
        matches_sheet.append(
            [
                "",
                "No matching past RCAs met the configured threshold.",
                "",
                "",
                "",
                "",
                current_problem,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                f"Threshold: {round(min_score * 100)}%",
            ]
        )

    widths = [11, 18, 13, 18, 20, 24, 42, 38, 44, 36, 36, 36, 18, 24, 13, 14, 16, 42, 52]
    for idx, width in enumerate(widths, start=1):
        matches_sheet.column_dimensions[get_column_letter(idx)].width = width
    matches_sheet.freeze_panes = "A2"
    matches_sheet.auto_filter.ref = matches_sheet.dimensions
    for row in matches_sheet.iter_rows():
        for cell in row:
            cell.alignment = wrap
        if row[0].row > 1 and row[0].row % 2 == 0:
            for cell in row:
                cell.fill = pale_fill

    for cell in summary[1]:
        cell.fill = title_fill
        cell.font = white_font

    wb.save(output_path)
    return output_path


def append_memory_to_context(context: str | None, evidence_pack: str | None) -> str | None:
    if not evidence_pack:
        return context
    if context:
        return f"{context.strip()}\n\n{evidence_pack}"
    return evidence_pack


def generic_root_cause_issue(root_cause: str) -> str | None:
    normalized = " ".join(TOKEN_RE.findall(root_cause.lower()))
    if normalized in GENERIC_ROOT_CAUSES:
        return "root_cause is too generic; name the concrete failed control, component, config, index, pool, route, secret, scheduler, schema, or alert rule."
    if len(_tokens(root_cause)) <= 3 and any(word in normalized for word in ("failure", "issue", "problem", "gap")):
        return "root_cause is too short and generic; include the specific mechanism."
    return None
